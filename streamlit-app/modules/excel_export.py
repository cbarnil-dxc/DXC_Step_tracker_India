"""
Excel export module for DXC Step Tracker.
Contains functions for generating comprehensive Excel exports.
"""

import io
import json
import logging
import streamlit as st
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .db import supabase


def generate_comprehensive_export():
    """Generate comprehensive Excel export with user and team views."""
    try:
        # Fetch all users with their forms
        users = supabase.table("users").select("*").execute().data
        forms = supabase.table("forms").select("*").execute().data
        teams = supabase.table("teams").select("*").execute().data
        
        # Load challenges data
        challenges_path = Path(__file__).resolve().parent.parent / ".streamlit" / "static" / "assets" / "Challenges.json"
        with open(challenges_path, "r", encoding="utf-8") as f:
            challenges_data = json.load(f)
        
        # Create workbook
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Users"
        ws_teams = wb.create_sheet("Teams")
        ws_stats = wb.create_sheet("Statistics")
        
        # Define DXC blue color
        dxc_blue = "7BA4DB"
        header_fill = PatternFill(start_color=dxc_blue, end_color=dxc_blue, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== USERS SHEET ====================
        user_headers = [
            "Name", "Email", "Team Name", "Team Leader",
            "Total Submissions", "Total Steps", "Average Daily Steps",
            "Max Steps (Single Day)", "First Submission Date", "Last Submission Date",
            "Challenges Completed"
        ]
        
        # Write headers
        for col_num, header in enumerate(user_headers, 1):
            cell = ws_users.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Calculate user statistics
        user_stats = []
        for user in users:
            user_forms = [f for f in forms if f["user_id"] == user["user_id"]]
            total_submissions = len(user_forms)
            total_steps = sum(f["form_stepcount"] for f in user_forms)
            avg_daily = total_steps / total_submissions if total_submissions > 0 else 0
            max_steps = max((f["form_stepcount"] for f in user_forms), default=0)
            
            dates = [f["form_date"] for f in user_forms if f["form_date"]]
            first_date = min(dates) if dates else "N/A"
            last_date = max(dates) if dates else "N/A"
            
            # Get team info
            team_name = "No Team"
            team_leader = "N/A"
            if user.get("team_id"):
                team = next((t for t in teams if t["team_id"] == user["team_id"]), None)
                if team:
                    team_name = team["team_name"]
                    # Get team leader name
                    leader = next((u for u in users if u["user_id"] == team["team_leader_id"]), None)
                    team_leader = leader["user_name"] if leader else "N/A"
            
            # Get completed challenges
            completed_challenges = []
            for challenge_key, challenge_data in challenges_data.items():
                if challenge_key in user_forms:
                    completed_challenges.append(challenge_data["title"])
            challenges_str = ", ".join(completed_challenges) if completed_challenges else "None"
            
            user_stats.append({
                "Name": user["user_name"],
                "Email": user["user_email"],
                "Team Name": team_name,
                "Team Leader": team_leader,
                "Total Submissions": total_submissions,
                "Total Steps": total_steps,
                "Average Daily Steps": round(avg_daily, 2),
                "Max Steps (Single Day)": max_steps,
                "First Submission Date": first_date,
                "Last Submission Date": last_date,
                "Challenges Completed": challenges_str
            })
        
        # Sort by total steps (descending)
        user_stats.sort(key=lambda x: x["Total Steps"], reverse=True)
        
        # Write user data
        for row_num, stat in enumerate(user_stats, 2):
            for col_num, header in enumerate(user_headers, 1):
                cell = ws_users.cell(row=row_num, column=col_num)
                cell.value = stat[header]
                cell.border = border_style
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for col_num, header in enumerate(user_headers, 1):
            max_length = len(header)
            for row_num in range(2, len(user_stats) + 2):
                cell_value = ws_users.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_users.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # ==================== TEAMS SHEET ====================
        team_headers = [
            "Team Name", "Leader", "Member Count", "Total Team Steps",
            "Average Steps per Member", "Total Submissions", "Average Daily Steps per Member",
            "Max Steps (Single Day)", "Members"
        ]
        
        # Write headers
        for col_num, header in enumerate(team_headers, 1):
            cell = ws_teams.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Calculate team statistics
        team_stats = []
        for team in teams:
            team_users = [u for u in users if u.get("team_id") == team["team_id"]]
            member_count = len(team_users)
            
            # Get team leader name
            leader = next((u for u in users if u["user_id"] == team["team_leader_id"]), None)
            leader_name = leader["user_name"] if leader else "N/A"
            
            # Calculate team statistics
            team_forms = [f for f in forms if f["user_id"] in [u["user_id"] for u in team_users]]
            total_team_steps = sum(f["form_stepcount"] for f in team_forms)
            avg_steps_per_member = total_team_steps / member_count if member_count > 0 else 0
            total_submissions = len(team_forms)
            avg_daily_per_member = total_team_steps / member_count if member_count > 0 else 0
            max_steps = max((f["form_stepcount"] for f in team_forms), default=0)
            
            # Get member names
            member_names = ", ".join([u["user_name"] for u in team_users])
            
            team_stats.append({
                "Team Name": team["team_name"],
                "Leader": leader_name,
                "Member Count": member_count,
                "Total Team Steps": total_team_steps,
                "Average Steps per Member": round(avg_steps_per_member, 2),
                "Total Submissions": total_submissions,
                "Average Daily Steps per Member": round(avg_daily_per_member, 2),
                "Max Steps (Single Day)": max_steps,
                "Members": member_names
            })
        
        # Sort by total team steps (descending)
        team_stats.sort(key=lambda x: x["Total Team Steps"], reverse=True)
        
        # Write team data
        for row_num, stat in enumerate(team_stats, 2):
            for col_num, header in enumerate(team_headers, 1):
                cell = ws_teams.cell(row=row_num, column=col_num)
                cell.value = stat[header]
                cell.border = border_style
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for col_num, header in enumerate(team_headers, 1):
            max_length = len(header)
            for row_num in range(2, len(team_stats) + 2):
                cell_value = ws_teams.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_teams.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # ==================== STATISTICS SHEET ====================
        stats_headers = ["Metric", "Value"]
        
        # Write headers
        for col_num, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        # Calculate general statistics
        total_submissions = len(forms)
        unique_steppers = len(set(f["user_id"] for f in forms))
        total_steps = sum(f["form_stepcount"] for f in forms)
        avg_steps_per_submission = total_steps / total_submissions if total_submissions > 0 else 0
        
        # Find most steps in a single day
        daily_steps = {}
        for f in forms:
            date = f.get("form_date", "Unknown")
            if date not in daily_steps:
                daily_steps[date] = 0
            daily_steps[date] += f["form_stepcount"]
        
        max_daily_steps = max(daily_steps.values()) if daily_steps else 0
        max_day = max(daily_steps, key=daily_steps.get) if daily_steps else "N/A"
        
        # Date range
        dates = [f.get("form_date") for f in forms if f.get("form_date")]
        first_submission_date = min(dates) if dates else "N/A"
        last_submission_date = max(dates) if dates else "N/A"
        
        # Verification status
        verified_count = sum(1 for f in forms if f.get("form_verified", False))
        unverified_count = total_submissions - verified_count
        
        # Challenge completions
        challenge_completions = sum(1 for f in forms if f.get("form_filepath") and f.get("form_filepath").startswith("challenge_"))
        
        # High step submissions (>20,000)
        high_step_count = sum(1 for f in forms if f.get("form_stepcount", 0) > 20000)
        
        # Build statistics data
        statistics = [
            {"Metric": "Total Submissions", "Value": total_submissions},
            {"Metric": "Number of Steppers (Unique Users)", "Value": unique_steppers},
            {"Metric": "Total Steps", "Value": total_steps},
            {"Metric": "Average Steps per Submission", "Value": round(avg_steps_per_submission, 2)},
            {"Metric": "Most Steps in a Single Day", "Value": max_daily_steps},
            {"Metric": "Date of Most Steps", "Value": max_day},
            {"Metric": "First Submission Date", "Value": first_submission_date},
            {"Metric": "Last Submission Date", "Value": last_submission_date},
            {"Metric": "Verified Submissions", "Value": verified_count},
            {"Metric": "Unverified Submissions", "Value": unverified_count},
            {"Metric": "Challenge Completions", "Value": challenge_completions},
            {"Metric": "High Step Submissions (>20,000)", "Value": high_step_count},
            {"Metric": "Total Teams", "Value": len(teams)},
            {"Metric": "Total Users Registered", "Value": len(users)}
        ]
        
        # Write statistics data
        for row_num, stat in enumerate(statistics, 2):
            cell = ws_stats.cell(row=row_num, column=1)
            cell.value = stat["Metric"]
            cell.border = border_style
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            cell = ws_stats.cell(row=row_num, column=2)
            cell.value = stat["Value"]
            cell.border = border_style
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto-adjust column widths
        for col_num, header in enumerate(stats_headers, 1):
            max_length = len(header)
            for row_num in range(2, len(statistics) + 2):
                cell_value = ws_stats.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_stats.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    except Exception:
        logging.error("Error generating comprehensive export")
        st.error("Error generating report.")
        return None
